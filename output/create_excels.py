"""
SHUWAN 動的Excelジェネレーター
日次PDCAで更新されたcontextファイルから最新データを読み込み、
Excelレポートを自動再生成する

実行: python ~/.claude/output/create_excels.py
"""
import re
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== スタイル定義 =====
HDR_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
HDR_FILL  = PatternFill('solid', fgColor='2F5496')
DATA_FONT = Font(name='Arial', size=10)
BLUE_FONT = Font(name='Arial', size=10, color='0070C0')
BOLD_FONT = Font(name='Arial', bold=True, size=10)
GREEN_FONT = Font(name='Arial', bold=True, size=10, color='375623')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),  bottom=Side(style='thin', color='D9D9D9'))
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

BASE   = 'C:/Users/hara/.claude'
OUTPUT = BASE + '/output'
TODAY  = date.today().strftime('%Y-%m-%d')

# ===== ユーティリティ =====
def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font, cell.fill, cell.alignment, cell.border = HDR_FONT, HDR_FILL, CENTER, THIN_BORDER

def style_data(ws, row, cols, font=None, align=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font      = font or DATA_FONT
        cell.border    = THIN_BORDER
        cell.alignment = align or (CENTER if c > 1 else LEFT)

def auto_width(ws, cols, mn=14, mx=38):
    for c in range(1, cols+1):
        ws.column_dimensions[get_column_letter(c)].width = min(mx, max(mn, 20))

def read_file(rel_path):
    try:
        with open(BASE + '/' + rel_path, encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        print(f'  [warn] {rel_path}: {e}')
        return ''

def parse_md_table(text, header_re):
    """Markdownテーブルをリスト of リストで返す（ヘッダー行は除外）"""
    lines = text.split('\n')
    rows, capturing = [], False
    for line in lines:
        if re.search(header_re, line, re.IGNORECASE):
            capturing = True
            continue
        if capturing:
            if re.match(r'\s*\|[-:\s|]+\|', line):
                continue
            if line.startswith('|'):
                cells = [c.strip() for c in line.split('|')[1:-1]]
                rows.append(cells)
            elif rows:
                break
    return rows

def int_val(s):
    """'1,234万円' → int"""
    s = s.replace(',', '').replace('円', '').replace('碗', '').replace('—', '0').strip()
    if '万' in s:
        s = s.replace('万', '')
        try: return int(float(s) * 10000)
        except: return 0
    try: return int(s)
    except: return 0

def float_pct(s):
    """'20%' → 0.20"""
    s = s.replace('%', '').strip()
    try: return float(s) / 100
    except: return 0.0

# ===== データ読み込み =====

def load_sales():
    text = read_file('context/memos/shuwan-sales-summary.md')
    rows = parse_md_table(text, r'月.*売上.*税抜.*販売数')
    data = []
    for r in rows:
        if len(r) < 3 or not r[0].strip(): continue
        try:
            month  = r[0].strip()
            sales  = int_val(r[1])
            units  = int_val(r[2])
            margin = int_val(r[3]) if len(r) > 3 else int(sales * 0.6)
            if sales > 0:
                data.append([month, sales, units, margin])
        except: pass
    # フォールバック
    if not data:
        data = [
            ['2025/01',126000,35,76290], ['2025/02',108000,32,64894],
            ['2025/03',104400,34,60940], ['2025/04',295920,88,197016],
            ['2025/05',313552,101,225910],['2025/06',1056880,381,670576],
            ['2025/07',1157305,410,723570],['2025/08',1779930,656,1299544],
            ['2025/09',1158242,421,729584],['2025/10',773560,285,398585],
            ['2025/11',875236,297,480413],['2025/12',1942194,666,1252092],
            ['2026/01',1915894,700,997454],['2026/02',1624394,551,884723],
        ]
    return data

def load_channels():
    text = read_file('context/strategy/kpi-targets.md')
    rows = parse_md_table(text, r'チャネル.*年間目標.*月平均')
    data = []
    for r in rows:
        if len(r) < 4 or '合計' in r[0] or '**' not in r[0] and r[0].strip() == '': continue
        name = r[0].strip().replace('**','').replace('（Shopify）','（Shopify）')
        if not name or '合計' in name: continue
        try:
            target  = int_val(r[1])
            monthly_s = r[2].replace('万円','0000').replace(',','').replace('—','0').replace('円','').strip()
            monthly = int(monthly_s) if monthly_s.isdigit() else 0
            ratio   = float_pct(r[3])
            price   = int_val(r[4]) if len(r) > 4 else 3000
            units   = int_val(r[5]) if len(r) > 5 else 0
            if target > 0:
                data.append([name, target, monthly, ratio, price, units])
        except: pass
    if not data:
        data = [
            ['自社EC（Shopify）',20000000,1670000,0.20,3900,428],
            ['Amazon',          15000000,1250000,0.15,4200,298],
            ['外販（エルスタイル等）',25000000,2080000,0.25,2700,771],
            ['酒販店（卸売）',   15000000,1250000,0.15,2800,446],
            ['MAKUAKE（年2回）', 10000000,0,      0.10,3500,0  ],
            ['料飲店',           15000000,1250000,0.15,2600,481],
        ]
    return data

def load_quarters():
    text = read_file('context/strategy/kpi-targets.md')
    rows = parse_md_table(text, r'四半期.*期間.*売上目標')
    data = []
    for r in rows:
        if len(r) < 4: continue
        try:
            q     = r[0].strip().replace('**','')
            period = r[1].strip()
            target = int_val(r[2])
            cumul  = int_val(r[3])
            rate   = float_pct(r[4]) if len(r) > 4 else 0
            basis  = r[5].strip() if len(r) > 5 else ''
            if q and target > 0:
                data.append([q, period, target, cumul, rate, basis])
        except: pass
    if not data:
        data = [
            ['Q1','3-5月', 15000000,15000000,0.15,'基盤構築期'],
            ['Q2','6-8月', 25000000,40000000,0.40,'父の日+MAKUAKE①'],
            ['Q3','9-11月',25000000,65000000,0.65,'MAKUAKE②+秋需要'],
            ['Q4','12-1月',35000000,100000000,1.00,'年末ギフト最大化'],
        ]
    return data

def load_competitors():
    text = read_file('context/market/competitors.md')
    comps = []
    sections = re.findall(r'### \d+\.\s+(.+?)\n(.*?)(?=### \d+\.|## |\Z)', text, re.DOTALL)
    for name, body in sections:
        name = name.strip()
        def g(pattern): m = re.search(pattern, body); return m.group(1).strip() if m else '-'
        comps.append([
            name,
            g(r'\*\*価格帯\*\*:\s*(.+)'),
            g(r'\*\*特徴\*\*:\s*(.+)'),
            g(r'\*\*チャネル\*\*:\s*(.+)'),
            g(r'\*\*強み\*\*:\s*(.+)'),
            g(r'\*\*弱み\*\*:\s*(.+)'),
            '×',
        ])
    shuwan = ['SHUWAN（自社）','3,900〜18,000円','磁器・日本酒専用設計・意匠登録済み',
              '自社EC/Amazon/酒販店/料飲店','日本酒専用×科学的根拠×意匠登録','認知度向上中','○']
    comps = [shuwan] + comps
    if len(comps) < 2:
        comps += [
            ['うすはりグラス','1,100〜3,300円','極薄ガラス','百貨店/Amazon','知名度高','割れやすい','×'],
            ['能作','5,000〜15,000円','錫100%','直営店/百貨店','ブランド力','高価格','×'],
            ['田島硝子','5,000〜6,000円','富士山デザイン','Amazon/楽天','ギフト人気','汎用品','×'],
            ['木本硝子','3,000〜8,000円','香り設計','自社EC/百貨店','日本酒専用設計','認知度低','○'],
        ]
    return comps

def load_customers():
    text = read_file('context/sales/customer-insights.md')
    rows = []
    for category, pattern in [('酒販店', r'### 酒販店\n(.*?)(?=###|##|\Z)'),
                               ('料飲店', r'### 料飲店\n(.*?)(?=###|##|\Z)'),
                               ('外販',   r'### 外販\n(.*?)(?=###|##|\Z)')]:
        m = re.search(pattern, text, re.DOTALL)
        if m:
            names = re.split(r'[、,\n・\-]+', m.group(1))
            for n in names:
                n = n.strip()
                if len(n) > 1 and not n.startswith('#'):
                    rows.append([category, n, ''])
    if not rows:
        rows = [
            ['酒販店','久楽屋',''],['酒販店','千鳥屋',''],['酒販店','鍵や',''],
            ['酒販店','五十嵐酒店',''],['料飲店','YAKITORI燃',''],
            ['外販','エルスタイル',''],['外販','ブラックス',''],
        ]
    return rows

def load_pipeline():
    text = read_file('context/sales/customer-insights.md')
    rows = []
    # Gmail/Slack取り込みセクションから商談情報を抽出
    section = re.search(r'### 新規商談.*?\n(.*?)(?=###|## |\Z)', text, re.DOTALL)
    if section:
        items = re.findall(r'\d+\.\s+\*\*(.+?)\*\*.*?[:：]\s*(.+)', section.group(1))
        for name, note in items:
            rows.append([name.strip(), note.strip()[:50], TODAY, '確認中'])
    # デフォルト追加
    defaults = [
        ['ヴィノスやまざき（戸塚）','しゅわんグラス前向き検討中','2026-03-25','見積り提出フォロー'],
        ['篠澤酒舗','しゅわんグラス案内送付済み','2026-03-25','返答待ち'],
        ['酒の勝鬨','しゅわんグラス案内送付済み','2026-03-25','返答待ち'],
        ['桑原商店','見積り送付済み','2026-03-25','発注確認'],
        ['いまでや×Sakesuki','海外輸出三者商談実施','2026-03-26','MOQ・輸出価格・梱包仕様の確認依頼'],
    ]
    existing_names = {r[0] for r in rows}
    for d in defaults:
        if d[0] not in existing_names:
            rows.append(d)
    return rows

# ===== Excel生成関数 =====

def sheet_meta(wb):
    """更新情報シートを追加"""
    ws = wb.create_sheet('更新情報')
    ws.append(['項目', '内容'])
    style_header(ws, 1, 2)
    ws.append(['生成日', TODAY])
    ws.append(['生成方法', '日次PDCA自動生成（create_excels.py）'])
    ws.append(['データソース', '~/.claude/context/ 配下の最新MDファイル'])
    for r in range(2, 5):
        style_data(ws, r, 2)
    auto_width(ws, 2, 20, 50)

def create_cmo_excel():
    wb = Workbook()

    # --- KPI目標 ---
    ws = wb.active
    ws.title = 'KPI目標'
    headers = ['チャネル','年間目標','月平均','構成比','想定単価','月間碗数']
    ws.append(headers); style_header(ws, 1, 6)
    channels = load_channels()
    for i, row in enumerate(channels):
        ws.append(row); style_data(ws, i+2, 6, BLUE_FONT)
    tr = len(channels) + 2
    ws.append(['合計', f'=SUM(B2:B{tr-1})', f'=SUM(C2:C{tr-1})',
               f'=SUM(D2:D{tr-1})', '約3,000円', f'=SUM(F2:F{tr-1})'])
    style_data(ws, tr, 6, BOLD_FONT)
    for r in range(2, tr+1):
        ws.cell(r,2).number_format='#,##0'; ws.cell(r,3).number_format='#,##0'
        ws.cell(r,4).number_format='0%';    ws.cell(r,5).number_format='#,##0'
        ws.cell(r,6).number_format='#,##0'
    auto_width(ws, 6)

    # --- 四半期マイルストーン ---
    ws2 = wb.create_sheet('四半期マイルストーン')
    ws2.append(['四半期','期間','売上目標','累計','達成率','根拠']); style_header(ws2, 1, 6)
    for i, row in enumerate(load_quarters()):
        ws2.append(row); style_data(ws2, i+2, 6, BLUE_FONT)
        ws2.cell(i+2,3).number_format='#,##0'; ws2.cell(i+2,4).number_format='#,##0'
        ws2.cell(i+2,5).number_format='0%'
    auto_width(ws2, 6)

    # --- 予算配分 ---
    ws3 = wb.create_sheet('予算配分')
    ws3.append(['項目','年間予算','備考']); style_header(ws3, 1, 3)
    budget = [
        ['広告費（Google/Amazon）',6000000,'ROAS400%目標'],
        ['YouTube制作費',1000000,'月4本×12ヶ月'],
        ['MAKUAKE手数料（2回）',2000000,'各回20%×1,000万円目標'],
        ['PR・プレスリリース',800000,'月1件掲載目標'],
        ['展示会・イベント',500000,'年2〜4回参加'],
    ]
    for i, row in enumerate(budget):
        ws3.append(row); style_data(ws3, i+2, 3, BLUE_FONT)
        ws3.cell(i+2,2).number_format='#,##0'
    tr3 = len(budget) + 2
    ws3.append(['合計', f'=SUM(B2:B{tr3-1})', '売上の約10%'])
    style_data(ws3, tr3, 3, BOLD_FONT); ws3.cell(tr3,2).number_format='#,##0'
    auto_width(ws3, 3, 22)

    sheet_meta(wb)
    path = OUTPUT + '/cmo/SHUWAN_KPI_年間計画.xlsx'
    wb.save(path); print(f'  CMO: {path}')

def create_researcher_excel():
    wb = Workbook()

    # --- 競合比較 ---
    ws = wb.active
    ws.title = '競合比較'
    headers = ['ブランド','価格帯','特徴・素材','チャネル','強み','弱み','日本酒専用']
    ws.append(headers); style_header(ws, 1, 7)
    for i, row in enumerate(load_competitors()):
        ws.append(row)
        style_data(ws, i+2, 7, GREEN_FONT if i == 0 else BLUE_FONT)
    auto_width(ws, 7, 14, 30)

    # --- 月次売上実績 ---
    ws2 = wb.create_sheet('月次売上実績')
    ws2.append(['月','売上(税抜)','販売数(碗)','限界利益','前月比','累計売上'])
    style_header(ws2, 1, 6)
    sales = load_sales(); cumul = 0
    for i, row in enumerate(sales):
        cumul += row[1]; ws2.append(row[:4] + ['', cumul])
        style_data(ws2, i+2, 6, BLUE_FONT)
        ws2.cell(i+2,2).number_format='#,##0'; ws2.cell(i+2,4).number_format='#,##0'
        ws2.cell(i+2,6).number_format='#,##0'
        if i > 0:
            ws2.cell(i+2,5).value = f'=B{i+2}/B{i+1}-1'
            ws2.cell(i+2,5).number_format='0%'
    tr = len(sales) + 2
    ws2.append(['合計', f'=SUM(B2:B{tr-1})', f'=SUM(C2:C{tr-1})', f'=SUM(D2:D{tr-1})','',''])
    style_data(ws2, tr, 6, BOLD_FONT)
    ws2.cell(tr,2).number_format='#,##0'; ws2.cell(tr,4).number_format='#,##0'
    auto_width(ws2, 6)

    # --- KPI達成状況 ---
    ws3 = wb.create_sheet('KPI達成状況')
    ws3.append(['KPI指標','年間目標','最新月実績','達成率','必要月平均','現状ペース(年換算)'])
    style_header(ws3, 1, 6)
    latest = load_sales()[-1] if load_sales() else ['-',0,0,0]
    kpi = [
        ['売上',    100000000, latest[1], f'=C2/B2', 8333333, f'=C2*12'],
        ['販売数(碗)',33300,   latest[2], f'=C3/B3', 2775,    f'=C3*12'],
        ['限界利益', 60000000, latest[3], f'=C4/B4', 5000000, f'=C4*12'],
    ]
    for i, row in enumerate(kpi):
        ws3.append(row); style_data(ws3, i+2, 6, BLUE_FONT)
        for c in [2,3,5,6]: ws3.cell(i+2,c).number_format='#,##0'
        ws3.cell(i+2,4).number_format='0.0%'
    auto_width(ws3, 6)

    sheet_meta(wb)
    path = OUTPUT + '/researcher/SHUWAN_市場競合分析.xlsx'
    wb.save(path); print(f'  Researcher: {path}')

def create_comm_planner_excel():
    wb = Workbook()

    # --- メディアミックス ---
    ws = wb.active
    ws.title = 'メディアミックス'
    headers = ['チャネル','役割','月額予算','年間予算','優先度','KPI']
    ws.append(headers); style_header(ws, 1, 6)
    media = [
        ['YouTube',          '認知・教育',   75000, 900000,  '高',  '登録5,000人'],
        ['Instagram',        '認知・転換',  150000,1800000,  '高',  'フォロワー1万人'],
        ['Googleリスティング','購買層捕捉', 200000,2400000,  '高',  'CPA2,000円以下'],
        ['Amazon広告',       '検索順位向上',150000,1800000,  '高',  'ACoS25%以下'],
        ['SEO',              '長期資産',    40000, 480000,   '中',  '月間5,000PV'],
        ['PR',               '信頼性付与',  75000, 900000,   '中高','メディア掲載月1件'],
    ]
    for i, row in enumerate(media):
        ws.append(row); style_data(ws, i+2, 6, BLUE_FONT)
        ws.cell(i+2,3).number_format='#,##0'; ws.cell(i+2,4).number_format='#,##0'
    tr = len(media)+2
    ws.append(['合計','',f'=SUM(C2:C{tr-1})',f'=SUM(D2:D{tr-1})','',''])
    style_data(ws, tr, 6, BOLD_FONT)
    ws.cell(tr,3).number_format='#,##0'; ws.cell(tr,4).number_format='#,##0'
    auto_width(ws, 6)

    # --- 季節キャンペーン ---
    ws2 = wb.create_sheet('季節キャンペーン')
    ws2.append(['月','テーマ','主要施策','重点チャネル']); style_header(ws2, 1, 4)
    campaigns = [
        ['4月','花見・春酒','桜×日本酒Instagram広告・しゅわんグラス訴求','Instagram'],
        ['5月','GW・MAKUAKE②','MAKUAKEローンチ・GWギフト訴求','MAKUAKE/全チャネル'],
        ['6月','父の日','父の日ギフトセット特集LP・桐箱セット','全チャネル'],
        ['8月','夏酒','冷酒×SHUWAN/しゅわんグラスコンテンツ','YouTube/Instagram'],
        ['9月','敬老の日','プレミアムラインギフト訴求','Amazon/自社EC'],
        ['12月','年末ギフト','全チャネル集中投資・干支モデル','全チャネル'],
        ['1月','お年賀','新年干支訴求・日本酒体験ギフト','自社EC'],
    ]
    for i, row in enumerate(campaigns):
        ws2.append(row); style_data(ws2, i+2, 4, BLUE_FONT)
    auto_width(ws2, 4)

    # --- CRM施策 ---
    ws3 = wb.create_sheet('CRM施策')
    ws3.append(['施策','タイミング','内容','KPI','優先度']); style_header(ws3, 1, 5)
    crm = [
        ['サンクスメール',   '購入直後',   '使い方ガイド+日本酒おすすめ','開封率50%','最高'],
        ['コンテンツメール', '購入3日後',  '日本酒の楽しみ方',           '開封率30%','高'],
        ['レビュー依頼',     '購入14日後', '特典付きレビュー依頼',       '投稿率10%','高'],
        ['追加購入提案',     '購入30日後', 'ギフト・セット提案',         'CVR5%',    '中'],
        ['季節案内',         '購入60日後', '新商品・季節限定案内',       'CVR3%',    '中'],
        ['メルマガ(コンテンツ)','毎月第1週','日本酒コンテンツ配信',     '開封率25%','中'],
        ['メルマガ(商品)',   '毎月第3週',  'キャンペーン告知',           'CTR3%',    '中'],
    ]
    for i, row in enumerate(crm):
        ws3.append(row); style_data(ws3, i+2, 5, BLUE_FONT)
    auto_width(ws3, 5)

    sheet_meta(wb)
    path = OUTPUT + '/communication-planner/SHUWAN_メディアプラン.xlsx'
    wb.save(path); print(f'  CommPlanner: {path}')

def create_cso_excel():
    wb = Workbook()

    # --- チャネル別営業戦略 ---
    ws = wb.active
    ws.title = 'チャネル別営業戦略'
    headers = ['チャネル','年間目標','月間DM数','返信率','成約数/月','卸値範囲','最低発注']
    ws.append(headers); style_header(ws, 1, 7)
    strategy = [
        ['酒販店',         '+50店','50件', '10%','4件','2,640〜2,880円','4碗'],
        ['料飲店',         '+60店','150件','5%', '5件','2,440〜2,950円','6碗'],
        ['外販',           '2,500万円','商談2件/月','—','1社/月','OEM対応','30碗'],
        ['しゅわんグラス卸','+20店','30件','15%','4件','1,820〜2,160円','6本'],
    ]
    for i, row in enumerate(strategy):
        ws.append(row); style_data(ws, i+2, 7, BLUE_FONT)
    auto_width(ws, 7)

    # --- 四半期別営業目標 ---
    ws2 = wb.create_sheet('四半期別営業目標')
    ws2.append(['四半期','酒販店新規','料飲店新規','外販売上','しゅわんグラス新規'])
    style_header(ws2, 1, 5)
    q_data = [
        ['Q1(3-5月)','+12店','+15店',5000000,'+5店'],
        ['Q2(6-8月)','+12店','+15店',6250000,'+5店'],
        ['Q3(9-11月)','+13店','+15店',6250000,'+5店'],
        ['Q4(12-1月)','+13店','+15店',7500000,'+5店'],
    ]
    for i, row in enumerate(q_data):
        ws2.append(row); style_data(ws2, i+2, 5, BLUE_FONT)
        ws2.cell(i+2,4).number_format='#,##0'
    tr2 = len(q_data)+2
    ws2.append(['合計','+50店','+60店',f'=SUM(D2:D{tr2-1})','+20店'])
    style_data(ws2, tr2, 5, BOLD_FONT); ws2.cell(tr2,4).number_format='#,##0'
    auto_width(ws2, 5)

    # --- 既存取引先 ---
    ws3 = wb.create_sheet('既存取引先')
    ws3.append(['区分','取引先名','状況・メモ']); style_header(ws3, 1, 3)
    for i, row in enumerate(load_customers()):
        ws3.append(row); style_data(ws3, i+2, 3)
    auto_width(ws3, 3, 18)

    # --- 新規商談パイプライン ---
    ws4 = wb.create_sheet('新規商談パイプライン')
    ws4.append(['商談先','状況・メモ','最終接触日','次アクション']); style_header(ws4, 1, 4)
    for i, row in enumerate(load_pipeline()):
        ws4.append(row[:4]); style_data(ws4, i+2, 4, BLUE_FONT)
    auto_width(ws4, 4, 18, 40)

    sheet_meta(wb)
    path = OUTPUT + '/cso/SHUWAN_営業計画.xlsx'
    wb.save(path); print(f'  CSO: {path}')

def create_cto_excel():
    wb = Workbook()

    # --- PDCA実行手順 ---
    ws = wb.active
    ws.title = 'PDCA実行手順'
    headers = ['ステップ','担当','タスク','依存関係','出力先','所要時間']
    ws.append(headers); style_header(ws, 1, 6)
    steps = [
        ['Step0', 'ユーザー',       '当月レベシェアExcelを格納',     'なし',       'context/memos/レベシェア実績/','5分'],
        ['Step1a','Researcher',     '市場・競合・口コミ調査更新',     'Step0完了後','context/market/',            '5分'],
        ['Step1b','CommunicationPlanner','コミュ施策の進捗評価・更新','Step0完了後','context/communication/',      '5分'],
        ['Step1c','CSO',            '営業進捗評価・パイプライン確認', 'Step0完了後','context/sales/',              '5分'],
        ['Step1d','CTO',            'プロンプト・技術改善点の洗い出し','Step0完了後','context/tech/',             '5分'],
        ['Step2', 'CMO',            'KPI統括判断・施策優先度の見直し','Step1完了後','context/strategy/',           '5分'],
        ['Step3', 'daily-pdca',     '日次MD生成・Excel再生成・報告', 'Step2完了後','output/',                     '3分'],
    ]
    for i, row in enumerate(steps):
        ws.append(row); style_data(ws, i+2, 6, BLUE_FONT)
    auto_width(ws, 6, 18, 40)

    # --- プロンプト一覧 ---
    ws2 = wb.create_sheet('プロンプト一覧')
    ws2.append(['#','プロンプト名','用途','対象エージェント','最終更新'])
    style_header(ws2, 1, 5)
    prompts = [
        ['1','月次PDCAレポート',       '実績分析・次月施策立案',         'CMO',               TODAY],
        ['2','Shopify商品ページ最適化', 'SEO・CVR改善',                  'CTO',               TODAY],
        ['3','Amazon出品最適化',       '検索順位向上・A+コンテンツ',     'CTO',               TODAY],
        ['4','SNSコンテンツ企画',      'YouTube/Instagram投稿計画',      'コミュプランナー',   TODAY],
        ['5','BtoB営業アプローチ文',   'DM・商談資料作成',               'CSO',               TODAY],
        ['6','MAKUAKEキャンペーン',    'CF企画・リターン設計',           'CMO/コミュプランナー',TODAY],
        ['7','90%が美味しい訴求コピー','EC商品ページ・広告コピー全般',   '全エージェント',     TODAY],
    ]
    for i, row in enumerate(prompts):
        ws2.append(row); style_data(ws2, i+2, 5, BLUE_FONT)
    auto_width(ws2, 5)

    # --- 技術課題トラッカー ---
    ws3 = wb.create_sheet('技術課題トラッカー')
    ws3.append(['課題','優先度','担当','期限','ステータス','備考']); style_header(ws3, 1, 6)
    issues = [
        ['GA4タグのShopify導入',                '最高','担当者未定','2026-04-15','未着手',   '全施策効果計測の前提'],
        ['Shopify「90%が美味しい」コピー反映',  '高',  '担当者未定','2026-03-28','未着手',   '最強訴求コピーの即時反映'],
        ['自社ECトップページ2商品並列表示',     '高',  '深田氏',   '2026-03-28','対応中',   'SHUWAN+しゅわんグラス並列'],
        ['Amazon A+コンテンツ追加',             '高',  'CTO',      '2026-04-07','計画中',   'B0DYD2DRKLブランドストーリー'],
        ['Amazon「意匠登録済み」表記追加',      '中',  '担当者未定','2026-04-15','未着手',   '差別化優位性の訴求'],
        ['Excel→MD自動更新スクリプト整備',     '中',  'CTO',      '2026-04-30','計画中',   '月次レベシェアExcelのMD変換自動化'],
    ]
    for i, row in enumerate(issues):
        ws3.append(row); style_data(ws3, i+2, 6, BLUE_FONT)
    auto_width(ws3, 6, 18, 40)

    sheet_meta(wb)
    path = OUTPUT + '/cto/SHUWAN_技術支援.xlsx'
    wb.save(path); print(f'  CTO: {path}')

# ===== エントリーポイント =====
if __name__ == '__main__':
    print(f'SHUWAN Excel生成開始: {TODAY}')
    create_cmo_excel()
    create_researcher_excel()
    create_comm_planner_excel()
    create_cso_excel()
    create_cto_excel()
    print('全ファイル生成完了')
